import pandas as pd
from openpyxl import load_workbook



def main():    
# Load the Excel file using openpyxl
    file_path = r"C:\Users\cryst\Desktop\personal_projects\post_mortem\data\rawdata\pplData.xlsx"
    wb = load_workbook(file_path)

    # Select the active sheet
    sheet = wb.active

    # Initialize an empty list to hold all rows of data
    data = []

    # Iterate through all rows and extract values (including empty ones)
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    # Convert the list of rows into a pandas DataFrame
    # The first row will be the column names (header)
    column_names = data[0]  # First row (index 0) is the column names
    data_rows = data[1:]    # All subsequent rows are the data

    # Create the DataFrame
    ppldata = pd.DataFrame(data_rows, columns=column_names)

    # Display the DataFrame to verify if all rows are loaded with correct column names
    print(ppldata.head())
    print(ppldata.columns)

    # Step 1: Forward fill to propagate missing values downwards for person-level info
    ppldata[['cbiNo', 'compName', 'inOut', 'Reason', 'Notes', 'Name', 'Link', 'PM Date', 
        'highestDeg', 'fieldHD', 'yearHD', 'yearFD', 'fieldFD', 'Alumni', 'Filename']] = ppldata[[
            'cbiNo', 'compName', 'inOut', 'Reason', 'Notes', 'Name', 'Link', 'PM Date', 
            'highestDeg', 'fieldHD', 'yearHD', 'yearFD', 'fieldFD', 'Alumni', 'Filename']].ffill()

    # Step 2: Create Master DataFrame with unique person information (first row is enough)
    master_df = ppldata[[
        'cbiNo', 'compName', 'inOut', 'Reason', 'Notes', 'Name', 'Link', 'PM Date', 
        'highestDeg', 'fieldHD', 'yearHD', 'yearFD', 'fieldFD', 'Alumni', 'Filename'
    ]].drop_duplicates(subset=['cbiNo'])

    # Step 3: Create Job DataFrame with job-related information
    job_df = ppldata[[
        'cbiNo', 'company', 'jobTitle', 'stMth', 'stYr', 'eMth', 'eYr'
    ]]

    master_df.to_pickle(r"C:\Users\cryst\Desktop\personal_projects\post_mortem\data\processed\master_df.pkl")
    job_df.to_pickle(r"C:\Users\cryst\Desktop\personal_projects\post_mortem\data\processed\job_df.pkl")
    
    
if __name__ == "__main__":
    main()